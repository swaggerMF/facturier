import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import os

file_path = 'C:/Users/simio/Desktop/facturi/Factura.xlsx'




# Function to modify the Excel sheet based on the input data and cell reference
def modify_excel(data_dict):
    # Load the Excel file using openpyxl
    wb = load_workbook(file_path)
    sheet = wb.active  # You can also specify a specific sheet if necessary

    cell_mapping = {
        "Nume client":(5,5),
        "Nr. factura":(0,5),
        "Registru comercial client":(7,6),
        "CIF client":(8,6),
        "Adresa client":(9,6),
        "Judet client":(11,6),
        "Tara client":(12,6)
    }

    for field, value in data_dict.items():
        if field in cell_mapping:
            row, col = cell_mapping[field]
            sheet.cell(row=row + 1, column=col + 1, value=value)

    # Generate a new Excel file
    nrfact = entry_nrfact.get()
    base, ext = os.path.splitext(file_path)
    new_file_path = f"{base} SI {nrfact}{ext}"
    wb.save(new_file_path)

    return new_file_path

# Function triggered by the button to open a file and modify it
def open_file_and_modify_with_mapping():
    data = {
        "Nr. factura":"SI " + entry_nrfact.get(),
        "Nume client":entry_numeclient.get(),
        "Registru comercial client": entry_regcomercial.get(),
        "CIF client": entry_cif.get(),
        "Adresa client":entry_adresa.get(),
        "Judet client":entry_jud.get(),
        "Tara client":entry_tara.get(),
        # Add more fields as needed
    }

    #file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        try:
            # Modify the Excel file and generate a new one
            new_file = modify_excel(data)
            messagebox.showinfo("Success", f"Fisierul a fost modificat: {new_file}")
        except Exception as e:
            messagebox.showerror("Error", f"A aparut o eroare: {e}")
        entry_nrfact.delete(0, tk.END)
        entry_numeclient.delete(0, tk.END)
        entry_regcomercial.delete(0, tk.END)
        entry_cif.delete(0, tk.END)
        entry_adresa.delete(0, tk.END)
        entry_jud.delete(0, tk.END)
        entry_tara.delete(0, tk.END)
# Create the GUI
root = tk.Tk()
root.geometry("700x500")
root.title("Facturier")

# Input label and field for data
tk.Label(root, text="Nr. factura:").pack(pady=5)
entry_nrfact = tk.Entry(root, width=30)
entry_nrfact.pack(pady=5)

tk.Label(root, text="Nume client:").pack(pady=5)
entry_numeclient = tk.Entry(root, width=30)
entry_numeclient.pack(pady=5)

tk.Label(root, text="Registru comercial client:").pack(pady=5)
entry_regcomercial = tk.Entry(root, width=30)
entry_regcomercial.pack(pady=5)

# Input label and field for cell reference (e.g., A1, B2)

tk.Label(root, text="CIF client:").pack(pady=5)
entry_cif = tk.Entry(root, width=30)
entry_cif.pack(pady=5)

tk.Label(root, text="Adresa client:").pack(pady=5)
entry_adresa = tk.Entry(root, width=30)
entry_adresa.pack(pady=5)

tk.Label(root, text="Judet client:").pack(pady=5)
entry_jud= tk.Entry(root, width=30)
entry_jud.pack(pady=5)

tk.Label(root, text="Tara client:").pack(pady=5)
entry_tara = tk.Entry(root, width=30)
entry_tara.pack(pady=5)

# Button to open file dialog and modify Excel
btn = tk.Button(root, text="Genereaza factura", command=open_file_and_modify_with_mapping)
btn.pack(pady=20)

# Run the GUI loop
root.mainloop()

# pyinstaller --onefile --noconsole facturier.py

